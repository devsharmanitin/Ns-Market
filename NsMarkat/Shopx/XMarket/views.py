import os
from django.contrib.auth.base_user import AbstractBaseUser
from .models import * 
from django.views import View
from django.conf import settings
from django.core.files import File
from django.shortcuts import render , redirect
import random , math
from django.contrib import messages
from django.contrib.auth import authenticate , login , logout , get_user_model
from .forms import *
from django.views.decorators.cache import never_cache
from django.http import JsonResponse
from django.db.models import Min , Max ,Sum
from django.contrib.auth.decorators import login_required





# Create your views here.


from decimal import Decimal

def Sections(request):
    PP = Product.objects.filter(Product_Section__isnull = True )
    SS = Section.objects.all()
    Count = SS.count()
    for i in PP:
        random_index = random.randint(0, Count-1)
        i.Product_Section = SS[random_index]
        print(i.Product_Section  )
        i.save()
    return render(request , 'main/Sec.html')

def REsave(request):
    Pp = Product.objects.all()
    for i in Pp:
        i.save()
    return redirect('HomePage')

def Pricee(request):
    Pp = Product.objects.all()
    for counter, product in enumerate(Pp):
    
        if "₹" in product.Price:
            price_str = product.Price.replace("₹", "").replace(",", "")
            product.Price = Decimal(price_str)
            product.save()
         
        if "%" in product.Discount:   
            price_str = product.Discount.replace("%", "").replace(",", "").replace(" OFF" , "").replace("-" , "")
            product.Discount = int(price_str)
            product.save()
            
        if ".00" in product.Price:
            price_str = product.Price.replace(".00", "")
            product.Price = int(price_str)
            product.save()

    return redirect('HomePage')


def AddItem(request):
    SliderImage = os.listdir('static/assets/img/slider/')
    for Item in SliderImage:
        if Item.endswith('.png') or Item.endswith('.jpg'):
            Image_Path = os.path.join('static/assets/img/slider/' , Item)
            with open(Image_Path , 'rb') as Image_File:
                
                GetModel = Slider()
                GetModel.SliderImage = Image_Path  
                GetModel.Sale = 10
                GetModel.Brand_Name = Item
                GetModel.save()
                
    return redirect('HomePage')


def HomePage(request):
    SLider_Images = Slider.objects.all()
    Banner_Images = BannerAera.objects.all()
    Main_Category = MainCategory.objects.all().order_by('-id')
    Product_Top_Deal = Product.objects.filter( Product_Section__name = 'Top Deals Of The Day' )
    Product_Top_Deals = Product_Top_Deal.order_by('id')[0:20]
    Top_Selling_Product = Product.objects.filter( Product_Section__name = 'Top Selling Products' )
    Top_Selling_Products = Top_Selling_Product.order_by('-id')[0:20]
    Top_Featured_Product = Product.objects.filter( Product_Section__name = 'Top Featured Products' )
    Top_Featured_Products = Top_Featured_Product.order_by('id')[0:25]
    Recommended_For_You = Product.objects.filter( Product_Section__name = 'Recommended For You' )
    Recommended_For_You = Recommended_For_You.order_by('-id')[0:25]
    
     
    Context = {
        'SLider_Images':SLider_Images,
        'Banner_Images':Banner_Images,
        'Main_Category':Main_Category,
        'Product_Top_Deals':Product_Top_Deals,
        'Top_Selling_Products':Top_Selling_Products,
        'Top_Featured_Products':Top_Featured_Products,
        'Recommended_For_You':Recommended_For_You,
    }
    return render(request ,'main/Home.html' ,Context)



def Product_Details(request , slug):
    GetProduct = Product.objects.filter(slug = slug)
    if GetProduct.exists():
        GetProduct = Product.objects.get(slug = slug)
    else:
        return redirect('Error404')
    Context = {
        'GetProduct':GetProduct,
    }
    return render(request , 'Products/ProductDetails.html' , Context)


def Error404(request):
    return render(request , 'Error/404.html')

def Authentication(request):
    return render(request , 'Authentication/MyAccount.html')

def Register(request):
    if request.method == "POST":
        Username = request.POST.get('Username')
        Email = request.POST.get('Email')
        Password = request.POST.get('Password')
        ConfPass = request.POST.get('ConfPassword')
        
        if not Password == ConfPass:
            messages.warning(request , 'Password Not Matching')
            return redirect('Register')
        if CustomUser.objects.filter( username = Username ).exists():
            messages.warning(request , 'Username Already Exists')
            return redirect('Register')
        if CustomUser.objects.filter( email = Email ).exists():
            messages.warning(request , 'Email Already Exists')
            
        NewUser = CustomUser.objects.create( username = Username , email = Email  )
        NewUser.set_password(Password)
        NewUser.save()
        return redirect('Register')
    return render(request , 'Authentication/MyAccount.html')
    
@never_cache 
def Login(request):
    if request.method == "POST":
        Username = request.POST.get('Username')
        Password = request.POST.get('Password')
        
        GetUser = authenticate(username = Username , password = Password)
        if GetUser is not None:
            login(request , GetUser)
            return redirect('HomePage')
        else:
            messages.warning(request , 'Credientials Not Matching') 
            return redirect('Login')
    return render(request , 'Authentication/MyAccount.html')
 
@never_cache   
def Logout(request):
    logout(request)
    return redirect('Login')


from django.contrib.auth.tokens import PasswordResetTokenGenerator
from django.contrib.sites.shortcuts import get_current_site  
from django.utils.encoding import force_str , force_bytes 
from django.utils.http import urlsafe_base64_encode, urlsafe_base64_decode  
from django.template.loader import render_to_string  
import six
from datetime import datetime,date
from django.core.mail import send_mail
from django.http import HttpResponse ,HttpResponseForbidden ,HttpResponseRedirect

def PasswordResetView(request):
    if request.method == "GET":
        
        FOrm = Password_Reset_form()
        Context = { 'form':FOrm }
        return render(request , 'password_reset_form.html' , Context)
    else:
        Form = Password_Reset_form(request.POST)
        if Form.is_valid():
            Email = Form.cleaned_data['email']
            GetUser = CustomUser.objects.filter(email = Email).exists()
            print(GetUser , "GGGGGGGGGGGGG")
            if GetUser is True:
                request.session['Password_Reset_Email'] = Email
                return redirect('Password_Reset_Done')
            else:
                messages.warning(request , 'Cannot Find User With Linked Email')
                return redirect('Reset_password')
    
    



class TokenGenerate(PasswordResetTokenGenerator):
        def _make_hash_value(self, user: CustomUser , timestamp: int):
            return (
                six.text_type(user.pk) + six.text_type(timestamp) + six.text_type(user.is_active)
            )
account_activate_token = TokenGenerate()
        
@never_cache       
def PasswordResetDoneView(request):
    Email = request.session.get('Password_Reset_Email')
    if Email:
        
        user = CustomUser.objects.get( email = Email )
        Current_Site = get_current_site(request)
        
        try:
            subject = 'Account Password Reset link to Reset your Account Password'
            message = render_to_string( 'Password_reset_Sent.html' , {
                'user' : user , 
                'domain' : Current_Site.domain , 
                'uid' : urlsafe_base64_encode(force_bytes(user.pk)),
                'token' : account_activate_token.make_token(user),
                'date' : date.today(),
            } )
            from_email = settings.EMAIL_HOST_USER
            recipient_list = [ Email ]
            
            send_mail(subject, message, from_email, recipient_list)
            return HttpResponse('Please confirm your email address to complete the registration')
        except Exception as Error:
            messages.warning(request , Error)
            return redirect('Reset_password')
    else:
        return redirect('Error404')
    
    
@never_cache   
def PasswordResetConfirmView(request , uidb64 , token):
    if request.method == "GET":
        UserModel = get_user_model()
        try:
            uid = urlsafe_base64_decode(uidb64 + '==')
            uid = force_str(uid)
            user = UserModel.objects.get(pk = uid)
        except (TypeError , ValueError , OverflowError , UserModel.DoesNotExist):
            user = None
        if user is not None and account_activate_token.check_token(user , token):
            user.is_active = True
            FOrm = Password_Confirm_form()
            Context = { 'form':FOrm }
            return render(request , 'password_reset_confirm.html' , Context)
        else:
            messages.warning(request , 'The password reset link was invalid, possibly because it has already been used. Please request a new password reset.')
            return redirect('Reset_password')
    else:
        Form = Password_Confirm_form(request.POST)
        if Form.is_valid():
            Password = Form.cleaned_data['password']
            Confirm_Password = Form.cleaned_data['Confirm_Password']
            if Password == Confirm_Password:
                user = Form.save(commit=False)
                user.set_password(Password)
                user.save()
                return redirect('Password_Reset_Complete')
 
@never_cache           
def PasswordResetCompleteView(request):
    if request.user.is_authenticated:
        logout(request)
    messages.info(request , 'Your New Password is Set')
    return render(request , 'password_reset_complete.html')


def UserProfile(request):
    user = CustomUser.objects.get(username = request.user) 
    orders = Order.objects.filter(user_id = user.id)
    Context = { 'Orders':orders }
    return render(request , 'Profile/profile.html' , Context)

import openpyxl                
def ExportExcel(request):
    wb = openpyxl.Workbook()
    sheet = wb.active
    
    fields = Product._meta.fields
    
    for index , field in enumerate(fields):
        sheet.cell(row=1 , column=index+1 , value = field.name.capitalize())
        
    PP = Product.objects.all()
    
    row = 2
    for pp in PP:
        for index , field in enumerate(fields):
            value = getattr(pp , field.attname)
            
            sheet.cell(row=row , column = index+1 , value=value)
        row +=1
        
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=Products.xlsx'
    wb.save(response)
    return response
        
        
        
        
def UpdateProfile(request):
    if request.method == "POST":
        First_Name = request.POST.get('First_name')
        last_name = request.POST.get('last_name')
        Number = request.POST.get('Number')
        Email = request.POST.get('email')
        
        GetUser = request.user
        GetUser.first_name = First_Name
        GetUser.last_name = last_name
        GetUser.number = Number
        GetUser.email = Email
        
        GetUser.save()
        return redirect('UserProfile')
    
def UserPersonnelDetails(request):
    if request.method == "POST":
        Gender = request.POST.get('Gender')
        Dob = request.POST.get('Dob')
        Address = request.POST.get('Address')
        
        GetUser = request.user
        GetUser.address = Address
        GetUser.dob = Dob
        GetUser.gender = Gender
        
        GetUser.save()
        return redirect('UserProfile')
    
    
def About(request):
    return render(request , 'main/About.html')

def ContactUs(request):
    return render(request , 'main/ContactUs.html')


def Products(request):
    GetProducts = None
    GetCategory = Category.objects.all()
    GetColors = Color.objects.all()
    GetSize = SIZE_CHOICES
    GetBrands = Brand.objects.all()
    
    MinPrice =  float('inf')
    MaxPrice = 0
    for i in Product.objects.all():
        Prices = int(i.Price)
        if Prices >= MaxPrice:
            MaxPrice = Prices
        if Prices < MinPrice:
            MinPrice = Prices
            
    filterPrice = request.GET.get("FilterPrice")
    ColorId = request.GET.get('ColorId')
    
    
    
    if filterPrice:
        GetProducts = Product.objects.filter(Price__lte = filterPrice)
        GetProducts = GetProducts.order_by('-Price')
    elif ColorId:
        GetProducts = Product.objects.filter(Product_Color__id = ColorId)
    
    else:
        GetProducts = Product.objects.all()
        
    
    Context = {
        'GetCategory':GetCategory,
        'GetProducts':GetProducts,
        'GetColors':GetColors,
        'GetSize':GetSize,
        'GetBrands':GetBrands,
        'MinPrice':MinPrice,
        'MaxPrice':MaxPrice,
        
    }
    return render(request , 'Products/Products.html' , Context)
        
    
def filter_data(request):
    categories = request.GET.getlist('GetCategory[]')
    SizeItem = request.GET.get('Size-item')
    Brand_item = request.GET.get('brand-item')

    allProducts = Product.objects.all().order_by('-id').distinct()
    if len(categories) > 0:
        allProducts = allProducts.filter(Product_Category__id__in=categories).distinct()
        
    if SizeItem:
        allProducts = Product.objects.filter(Product_Size = SizeItem)
        
    elif Brand_item:
        allProducts = Product.objects.filter(Product_Brand__id__in = Brand_item)
        print(allProducts)
        
    t = render_to_string('ajax/Product.html', {'GetProducts': allProducts ,  })

    return JsonResponse({'data': t})



from cart.cart import Cart 
        


@login_required()
def cart_add(request, id):
    cart = Cart(request)
    product = Product.objects.get(id=id)
    cart.add(product=product)
    return redirect("HomePage")


@login_required()
def item_clear(request, id):
    cart = Cart(request)
    product = Product.objects.get(id=id)
    cart.remove(product)
    return redirect("cart_detail")


@login_required(login_url='Account/Login')
def item_increment(request, id):
    cart = Cart(request)
    product = Product.objects.get(id=id)
    cart.add(product=product)
    return redirect("cart_detail")


@login_required(login_url='Account/Login')
def item_decrement(request, id):
    cart = Cart(request)
    product = Product.objects.get(id=id)
    cart.decrement(product=product)
    return redirect("cart_detail")


@login_required(login_url='Account/Login')
def cart_clear(request):
    cart = Cart(request)
    cart.clear()
    return redirect("cart_detail")

from cart.context_processor import cart_total_amount


def cart_detail(request):
    if request.user.is_authenticated:
        cart = request.session.get('cart')
        
        packing_cost = sum(i['packing_cost'] for i in cart.values() if i )
        tax = sum( i['tax'] for i in cart.values() if i)
        
        cart_items_total_amount = cart_total_amount(request)
        Total_Amount = int(cart_items_total_amount['cart_total_amount'])
        
        Coupan = None
        Valid_Coupan = None
        Invalid_COupan = None
        Discount_Code = None
        if request.method == "GET":
            CouponCode = request.GET.get('coupon_code')
            if CouponCode:
                try:
                    Coupan = Coupon_Code.objects.get(Code = CouponCode)
                    Valid_Coupan = "Are applicable on Current Coupan " 
                    Discount_Code = Total_Amount - (Total_Amount *(Coupan.Code_Discount / 100) ) 
                except:
                    Invalid_COupan = "Invalid Coupan "
            else:
                messages.warning(request , 'Invalid Code')
                Discount_Code = Total_Amount   
          
        Context = {
            'Packing_Cost':packing_cost,
            'Tax':tax,
            'Coupan':Coupan,
            'Valid_Coupan':Valid_Coupan,
            'Invalid_COupan':Invalid_COupan,
            'Discount_Code':Discount_Code,
        }
        return render(request, 'cart/cart_detail.html' , Context)
    else:
        return redirect('Login')
    
def CheckoutSuccess(request):
    return render(request , 'payment/success.html')


import stripe 
stripe.api_key = settings.STRIPE_SECRET_KEY

# def CheckOut(request):
#     cart_items_total_amount = cart_total_amount(request)
#     CartItem = request.session.get('cart')

#     if request.method == "POST":
#         token = request.POST['stripeToken']
#         try:
#             charge = stripe.Charge.create(
#                 amount = int(cart_items_total_amount),
#                 currency = 'usd',
#                 source = token,
#                 description = "Example chagre"
#             )
#             return redirect('cart_details')
#         except stripe.error.CardError as error:
#             error_msg = error.json_body['error']['message']
#             return render(request , 'payment/checkout.html', {'error':error_msg})
#     else:
#         return render(request ,'payment/checkout.html', {'cart_items_total_amount':cart_items_total_amount ,'publishable_key': settings.STRIPE_PUBLIC_KEY , 'id':id , 'amount':cart_items_total_amount  ,  })

    
stripe.api_key = settings.STRIPE_SECRET_KEY       
class CreateCheckoutSessionView(View):
    def post(self  , request , *args , **kwargs):
        Total_Amount = cart_total_amount(request)
        cart = request.session.get('cart')
        Domain = 'http://127.0.0.1:8000' 
        

        line_items_details = []
        for product_id , product_details in cart.items():
            item = {
                'price_data' : {
                    'currency' : 'inr' ,
                    'unit_amount' :  int(Total_Amount['cart_total_amount'])*100, 
                    'product_data' : {
                        'name' : product_details['Product_Name']
                    }
                },
                'quantity':1
            }   
            line_items_details.append(item)   
        success_url=Domain + '/products/checkout/success' ,
        cancel_url=Domain + '/products/item/checkout',
        try:
            checkout_session = stripe.checkout.Session.create(
            payment_method_types = ['card'],
            line_items =  line_items_details,
            mode = 'payment',
            success_url=Domain + '/products/checkout/success' ,
            cancel_url=Domain + '/products/item/checkout',
            ) 
            if success_url:
                
                for p_id , product_details in cart.items():
                    Product_id = Product.objects.get(id = p_id)
                    User_id = CustomUser.objects.get(id = request.user.id)
                    existing_order = Order.objects.filter(
                        user_id=User_id,
                        product_id=Product_id,
                    ).first()
                    if existing_order: 
                        if existing_order.quantity == None:
                            existing_order.quantity = 0
                            existing_order.quantity += 1
                        else:
                            existing_order.quantity += 1
                    else:
                        order = Order.objects.create(
                            user_id = User_id ,
                            product_id = Product_id , 
                            quantity = 1
                        )
                    order.save()
            return JsonResponse( {'id' : checkout_session.id , 'session_id': checkout_session.id  } )
        except Exception as error:
            print(error,"#############")
            messages.warning(request , error)
            return redirect('cart_detail')
        
      
def OrderView(request):
    user = CustomUser.objects.get(username = request.user) 
    orders = Order.objects.filter(user_id = user.id)
    Context = { 'Orders':orders }
    return render(request , 'payment/order.html' , Context)
     
        
class StripeIntentView(View):
    def get(self , request):
        payemnt_intent = stripe.PaymentIntent.create(
            amount = 1000,
            currency = 'inr',
        )
        return HttpResponse(payemnt_intent)
    
  
    
        
from django.db.models import Q     
def Search(request):
    if request.method == "POST":
        Search_item = request.POST.get('Search_item')
        
        GetProduct = Product.objects.filter(
            Q(Product_Name = Search_item) |
            Q(Price = Search_item) |
            Q(Product_Info = Search_item) | 
            Q(Model_Name = Search_item ) |
            Q(Product_Category = Search_item) | 
            Q(Tags = Search_item) |
            Q(Product_Desc = Search_item) |
            Q(Product_Section = Search_item) |
            Q(Product_Brand = Search_item)   
        )
        Context = {'Product':GetProduct}
        return render(request ,'main/Home.html' , Context)
    

class Wishlist(object):
    
    def __init__(self , request):
        self.request = request
        self.session = request.session
        wishlist = self.session.get(settings.WISHLIST_SESSION_ID)
        if not wishlist:
            # Save an empty wishlist in the session 
            wishlist = self.session[settings.WISHLIST_SESSION_ID] = {}
        self.wishlist = wishlist
        
    def add(self , product  , quantity=1 ):
        self.quantity = quantity
        id = product.id
        newItem = True
        if str(product.id) not in self.wishlist.keys():
            product.Price = int(product.Price)
            
            self.wishlist[product.id] = {
                'userid': self.request.user.id,
                'product_id': id,
                'Product_Name': product.Product_Name,
                'quantity': self.quantity,
                'Price': product.Price,
                'Product_Image': product.Product_Image , 
            }
        else:
            newItem = True
            messages.info(self.request , 'Item alredy exists in wishlist')
        self.save()
        
    def save(self):
        # update the session wishlist
        self.session[settings.WISHLIST_SESSION_ID] = self.wishlist
        # mark the session as "modified" to make sure it is saved
        self.session.modified = True
        
    def remove(self , product):
        
        # Remove item form wishlist 
        product_id = str(product.id)
        if product_id in self.wishlist:
            del self.wishlist[product_id]
            self.save()
            
    def wishlistclear(self):
        # empty Wihslist 
        self.session[settings.WISHLIST_SESSION_ID] = {}
        self.session.modified = True



def wishlist_add(request , id):
    wishlist = Wishlist(request)
    product = Product.objects.get(id = id)
    wishlist.add(product=product )
    return redirect('HomePage')

def wishlist_item_clear(request , id):
    wishlist = Wishlist(request)
    product = Product.objects.get(id=id)
    wishlist.remove(product)
    return redirect('Wishlist')

def wishlist_clear(request):
    wishlist = Wishlist(request)
    # wishlist.wishlistclear()
    return redirect('Wishlist')

def wishlist_detail(request):
    wishlist = request.session.get('wishlist')
    print(wishlist , "###############")
    return render(request , 'cart/wishlist.html')



def MarketBlogView(request):
    Blogs = MarketBlog.objects.all()
    Context = { 'Blogs':Blogs }
    return render(request , 'market/bloghome.html' , Context)


def BlogViewDeatils(request , id):
    Blog = MarketBlog.objects.get(id=id)
    Context = {'Blog':Blog}
    return render(request , 'market/blog_details.html' , Context)
    
    
    
        